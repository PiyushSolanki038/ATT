"""
excel_handler.py – Excel (openpyxl) and Google Sheets (gspread) I/O for
SISWIT Attendance Bot.

All Google Sheets writes run in a background thread guarded by an asyncio
Lock to prevent race conditions.
"""

import asyncio
import json as _json
import logging
from datetime import datetime
from pathlib import Path
from threading import Lock as ThreadLock

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

from config import (
    EXCEL_FILE,
    GOOGLE_SHEET_ID,
    GOOGLE_CREDS_FILE,
    GOOGLE_CREDS_JSON,
    ATTENDANCE_HEADERS,
    LEAVE_HEADERS,
    TIMEZONE,
)

logger = logging.getLogger(__name__)

# ── Concurrency primitives ───────────────────────────────────────────────────
_gs_lock = asyncio.Lock()
_excel_lock = ThreadLock()

# ── Styles ───────────────────────────────────────────────────────────────────
HEADER_FONT = Font(name="Calibri", bold=True, color="FFFFFF", size=11)
HEADER_FILL = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
HEADER_ALIGN = Alignment(horizontal="center", vertical="center", wrap_text=True)
THIN_BORDER = Border(
    left=Side(style="thin"),
    right=Side(style="thin"),
    top=Side(style="thin"),
    bottom=Side(style="thin"),
)
LATE_FILL = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
ONTIME_FILL = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")


# ══════════════════════════════════════════════════════════════════════════════
# EXCEL helpers
# ══════════════════════════════════════════════════════════════════════════════

def _get_or_create_workbook() -> Workbook:
    if Path(EXCEL_FILE).exists():
        try:
            return load_workbook(EXCEL_FILE)
        except Exception as exc:
            logger.warning("Corrupted workbook – recreating: %s", exc)
    return Workbook()


def _format_header(ws) -> None:
    """Apply professional header formatting to an attendance sheet."""
    for col_idx, header in enumerate(ATTENDANCE_HEADERS, 1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = HEADER_ALIGN
        cell.border = THIN_BORDER
    widths = [6, 12, 10, 20, 10, 15, 12, 50, 10, 18, 10]
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w


def _format_leave_header(ws) -> None:
    """Apply professional header formatting to the Leave Register sheet."""
    for col_idx, header in enumerate(LEAVE_HEADERS, 1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = HEADER_ALIGN
        cell.border = THIN_BORDER
    widths = [6, 12, 20, 15, 12, 30, 15, 10, 8, 12]
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w


def _get_next_sr_no(ws) -> int:
    max_sr = 0
    for row in ws.iter_rows(min_row=2, max_col=1, values_only=True):
        val = row[0]
        if val and isinstance(val, (int, float)):
            max_sr = max(max_sr, int(val))
    return max_sr + 1


def _update_dashboard(wb: Workbook) -> None:
    """Rebuild the Dashboard worksheet with per-month summary."""
    dash = "Dashboard"
    if dash in wb.sheetnames:
        ws = wb[dash]
    else:
        ws = wb.create_sheet(title=dash, index=0)

    for row in ws.iter_rows():
        for cell in row:
            cell.value = None

    ws.cell(row=1, column=1, value="SISWIT Attendance Dashboard")
    ws.cell(row=1, column=1).font = Font(name="Calibri", bold=True, size=14, color="4472C4")
    ws.cell(row=2, column=1, value=f"Last Updated: {datetime.now(TIMEZONE).strftime('%Y-%m-%d %H:%M:%S')}")

    for col, hdr in enumerate(["Month", "Total Entries"], 1):
        c = ws.cell(row=4, column=col, value=hdr)
        c.font = HEADER_FONT
        c.fill = HEADER_FILL

    row_idx = 5
    for sheet_name in wb.sheetnames:
        if sheet_name in (dash, "Leave Register"):
            continue
        sheet = wb[sheet_name]
        total = 0
        for r in range(2, sheet.max_row + 1):
            if sheet.cell(row=r, column=1).value:
                total += 1
        ws.cell(row=row_idx, column=1, value=sheet_name)
        ws.cell(row=row_idx, column=2, value=total)
        row_idx += 1

    ws.column_dimensions["A"].width = 20
    ws.column_dimensions["B"].width = 15


# ── Public Excel functions ───────────────────────────────────────────────────

def save_to_excel(
    emp_id: str,
    name: str,
    dept: str,
    date_str: str,
    day_name: str,
    submit_time: str,
    work: str,
    source: str,
    username: str,
    is_resubmission: bool = False,
) -> bool:
    """
    Persist an attendance record to the local Excel workbook.

    Creates the monthly sheet if needed, updates the Dashboard, and
    returns ``True`` on success.
    """
    with _excel_lock:
        try:
            wb = _get_or_create_workbook()
            dt = datetime.strptime(date_str, "%Y-%m-%d")
            sheet_name = dt.strftime("%B %Y")

            if sheet_name in wb.sheetnames:
                ws = wb[sheet_name]
            else:
                ws = wb.create_sheet(title=sheet_name)
                _format_header(ws)

            # Re-submission → update existing row in-place
            if is_resubmission:
                for r in range(2, ws.max_row + 1):
                    if (
                        ws.cell(row=r, column=5).value == emp_id
                        and ws.cell(row=r, column=2).value == date_str
                    ):
                        ws.cell(row=r, column=7, value=submit_time)
                        ws.cell(row=r, column=8, value=work)
                        ws.cell(row=r, column=11, value="Yes")
                        _update_dashboard(wb)
                        wb.save(EXCEL_FILE)
                        logger.info("Updated re-submission in Excel for %s on %s", emp_id, date_str)
                        return True
                # Existing row not found → will append below

            sr_no = _get_next_sr_no(ws)
            revision = "Yes" if is_resubmission else "No"
            row_data = [
                sr_no, date_str, day_name, name, emp_id, dept,
                submit_time, work, source, username, revision,
            ]

            next_row = ws.max_row + 1
            for col_idx, value in enumerate(row_data, 1):
                cell = ws.cell(row=next_row, column=col_idx, value=value)
                cell.border = THIN_BORDER
                cell.alignment = Alignment(horizontal="center", vertical="center")
                if col_idx == 8:  # Work Report – left-align
                    cell.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)

            _update_dashboard(wb)

            # Remove default "Sheet" if other sheets exist
            if "Sheet" in wb.sheetnames and len(wb.sheetnames) > 1:
                del wb["Sheet"]

            wb.save(EXCEL_FILE)
            logger.info("Saved to Excel: %s on %s", emp_id, date_str)
            return True

        except Exception as exc:
            logger.error("Excel save failed: %s", exc)
            return False


def update_entry_in_excel(emp_id: str, date_str: str, new_work: str) -> bool:
    """Update an existing entry's *Work Report* column in Excel."""
    with _excel_lock:
        try:
            if not Path(EXCEL_FILE).exists():
                return False
            wb = load_workbook(EXCEL_FILE)
            dt = datetime.strptime(date_str, "%Y-%m-%d")
            sheet_name = dt.strftime("%B %Y")
            if sheet_name not in wb.sheetnames:
                return False
            ws = wb[sheet_name]
            for r in range(2, ws.max_row + 1):
                if (
                    ws.cell(row=r, column=5).value == emp_id
                    and ws.cell(row=r, column=2).value == date_str
                ):
                    ws.cell(row=r, column=8, value=new_work)
                    ws.cell(row=r, column=11, value="Edited")
                    wb.save(EXCEL_FILE)
                    logger.info("Updated entry in Excel for %s on %s", emp_id, date_str)
                    return True
            return False
        except Exception as exc:
            logger.error("Excel update failed: %s", exc)
            return False


def remove_entry_from_excel(emp_id: str, date_str: str) -> bool:
    """Delete the row for *emp_id* on *date_str* from local Excel."""
    with _excel_lock:
        try:
            if not Path(EXCEL_FILE).exists():
                return False
            wb = load_workbook(EXCEL_FILE)
            dt = datetime.strptime(date_str, "%Y-%m-%d")
            sheet_name = dt.strftime("%B %Y")
            if sheet_name not in wb.sheetnames:
                return False
            ws = wb[sheet_name]
            for r in range(2, ws.max_row + 1):
                if (
                    ws.cell(row=r, column=5).value == emp_id
                    and ws.cell(row=r, column=2).value == date_str
                ):
                    ws.delete_rows(r)
                    _update_dashboard(wb)
                    wb.save(EXCEL_FILE)
                    logger.info("Removed entry from Excel for %s on %s", emp_id, date_str)
                    return True
            return False
        except Exception as exc:
            logger.error("Excel remove failed: %s", exc)
            return False


def save_leave_to_excel(
    emp_id: str,
    name: str,
    dept: str,
    leave_date: str,
    reason: str,
    approved_by: str,
    leave_count: int,
) -> bool:
    """Append a row to the *Leave Register* sheet in the local workbook."""
    with _excel_lock:
        try:
            wb = _get_or_create_workbook()
            lr_name = "Leave Register"
            if lr_name in wb.sheetnames:
                ws = wb[lr_name]
            else:
                ws = wb.create_sheet(title=lr_name)
                _format_leave_header(ws)

            sr_no = _get_next_sr_no(ws)
            deduction = ""
            if leave_count > 3:
                extra = leave_count - 3
                deduction = f"₹{extra * 500}"

            row_data = [
                sr_no, emp_id, name, dept, leave_date, reason,
                approved_by, "Approved", leave_count, deduction,
            ]
            next_row = ws.max_row + 1
            for col_idx, value in enumerate(row_data, 1):
                cell = ws.cell(row=next_row, column=col_idx, value=value)
                cell.border = THIN_BORDER
                cell.alignment = Alignment(horizontal="center", vertical="center")

            if "Sheet" in wb.sheetnames and len(wb.sheetnames) > 1:
                del wb["Sheet"]

            wb.save(EXCEL_FILE)
            logger.info("Saved leave to Excel: %s on %s", emp_id, leave_date)
            return True

        except Exception as exc:
            logger.error("Excel leave save failed: %s", exc)
            return False


# ══════════════════════════════════════════════════════════════════════════════
# GOOGLE SHEETS helpers
# ══════════════════════════════════════════════════════════════════════════════

def _get_gspread_client():
    """Authenticate and return a ``gspread.Client`` or ``None``."""
    try:
        import gspread
        from google.oauth2.service_account import Credentials

        scopes = [
            "https://www.googleapis.com/auth/spreadsheets",
            "https://www.googleapis.com/auth/drive",
        ]

        if GOOGLE_CREDS_JSON:
            creds_info = _json.loads(GOOGLE_CREDS_JSON)
            creds = Credentials.from_service_account_info(creds_info, scopes=scopes)
        elif Path(GOOGLE_CREDS_FILE).exists():
            creds = Credentials.from_service_account_file(GOOGLE_CREDS_FILE, scopes=scopes)
        else:
            logger.warning("No Google credentials found – Sheets integration disabled.")
            return None

        return gspread.authorize(creds)
    except Exception as exc:
        logger.error("Google Sheets auth failed: %s", exc)
        return None


def _ensure_sheet_headers(worksheet, headers: list[str]) -> None:
    """Self-heal headers: overwrite row 1 if it does not match *headers*."""
    try:
        existing = worksheet.row_values(1)
        if existing != headers:
            for i, h in enumerate(headers, 1):
                worksheet.update_cell(1, i, h)
            logger.info("Repaired headers in sheet '%s'", worksheet.title)
    except Exception as exc:
        logger.warning("Header check failed: %s", exc)


# ── Sync implementations (run inside worker thread) ─────────────────────────

def _save_to_google_sheets_sync(
    emp_id: str, name: str, dept: str, date_str: str,
    day_name: str, submit_time: str,
    work: str, source: str, username: str,
    is_resubmission: bool = False,
) -> None:
    if not GOOGLE_SHEET_ID:
        return
    client = _get_gspread_client()
    if not client:
        return
    try:
        spreadsheet = client.open_by_key(GOOGLE_SHEET_ID)
        dt = datetime.strptime(date_str, "%Y-%m-%d")
        sheet_name = dt.strftime("%B %Y")

        try:
            ws = spreadsheet.worksheet(sheet_name)
        except Exception:
            ws = spreadsheet.add_worksheet(title=sheet_name, rows=500, cols=len(ATTENDANCE_HEADERS))

        _ensure_sheet_headers(ws, ATTENDANCE_HEADERS)

        revision = "Yes" if is_resubmission else "No"

        # Re-submission → in-place update
        if is_resubmission:
            try:
                all_data = ws.get_all_values()
                for row_idx, row in enumerate(all_data[1:], 2):
                    if len(row) >= 5 and row[4] == emp_id and row[1] == date_str:
                        ws.update_cell(row_idx, 7, submit_time)
                        ws.update_cell(row_idx, 8, work)
                        ws.update_cell(row_idx, 11, "Yes")
                        logger.info("Updated re-submission in Sheets for %s on %s", emp_id, date_str)
                        return
            except Exception as exc:
                logger.warning("Re-submission in-place update failed – appending: %s", exc)

        all_data = ws.get_all_values()
        sr_no = len(all_data)  # header + existing rows gives next S.No.

        row_data = [
            sr_no, date_str, day_name, name, emp_id, dept,
            submit_time, work, source, username, revision,
        ]
        ws.append_row(row_data, value_input_option="USER_ENTERED")
        logger.info("Saved to Google Sheets: %s on %s", emp_id, date_str)

    except Exception as exc:
        logger.error("Google Sheets save failed: %s", exc)


def _update_entry_in_google_sheets_sync(emp_id: str, date_str: str, new_work: str) -> None:
    if not GOOGLE_SHEET_ID:
        return
    client = _get_gspread_client()
    if not client:
        return
    try:
        spreadsheet = client.open_by_key(GOOGLE_SHEET_ID)
        dt = datetime.strptime(date_str, "%Y-%m-%d")
        sheet_name = dt.strftime("%B %Y")
        try:
            ws = spreadsheet.worksheet(sheet_name)
        except Exception:
            return
        all_data = ws.get_all_values()
        for row_idx, row in enumerate(all_data[1:], 2):
            if len(row) >= 5 and row[4] == emp_id and row[1] == date_str:
                ws.update_cell(row_idx, 8, new_work)
                ws.update_cell(row_idx, 11, "Edited")
                logger.info("Updated entry in Google Sheets for %s on %s", emp_id, date_str)
                return
    except Exception as exc:
        logger.error("Google Sheets edit update failed: %s", exc)


def _save_leave_to_google_sheets_sync(
    emp_id: str, name: str, dept: str,
    leave_date: str, reason: str, approved_by: str,
    leave_count: int,
) -> None:
    if not GOOGLE_SHEET_ID:
        return
    client = _get_gspread_client()
    if not client:
        return
    try:
        spreadsheet = client.open_by_key(GOOGLE_SHEET_ID)
        lr_name = "Leave Register"
        try:
            ws = spreadsheet.worksheet(lr_name)
        except Exception:
            ws = spreadsheet.add_worksheet(title=lr_name, rows=500, cols=len(LEAVE_HEADERS))

        _ensure_sheet_headers(ws, LEAVE_HEADERS)

        all_data = ws.get_all_values()
        sr_no = len(all_data)

        deduction = ""
        if leave_count > 3:
            extra = leave_count - 3
            deduction = f"₹{extra * 500}"

        row_data = [
            sr_no, emp_id, name, dept, leave_date, reason,
            approved_by, "Approved", leave_count, deduction,
        ]
        ws.append_row(row_data, value_input_option="USER_ENTERED")
        logger.info("Saved leave to Google Sheets: %s on %s", emp_id, leave_date)

    except Exception as exc:
        logger.error("Google Sheets leave save failed: %s", exc)


def _load_attendance_from_google_sheets_sync() -> dict:
    """Load historical attendance from Google Sheets into a dict."""
    if not GOOGLE_SHEET_ID:
        return {}
    client = _get_gspread_client()
    if not client:
        return {}

    merged: dict = {}

    try:
        spreadsheet = client.open_by_key(GOOGLE_SHEET_ID)
        now_dt = datetime.now(TIMEZONE)

        sheets_to_load = [now_dt.strftime("%B %Y")]
        if now_dt.day <= 5:
            prev_month = now_dt.month - 1
            prev_year = now_dt.year
            if prev_month == 0:
                prev_month = 12
                prev_year -= 1
            prev_dt = now_dt.replace(year=prev_year, month=prev_month, day=1)
            sheets_to_load.append(prev_dt.strftime("%B %Y"))

        for sheet_name in sheets_to_load:
            try:
                ws = spreadsheet.worksheet(sheet_name)
                all_data = ws.get_all_values()
                for row in all_data[1:]:
                    if len(row) < 11:
                        continue
                    date_str = row[1]
                    emp_id = row[4]
                    if not date_str or not emp_id:
                        continue
                    merged.setdefault(date_str, {})
                    merged[date_str][emp_id] = {
                        "time": row[6],
                        "work": row[7],
                        "late": False,
                        "username": row[9] if len(row) > 9 else "",
                        "group": "Loaded",
                        "is_resubmission": (row[10].lower() == "yes") if len(row) > 10 else False,
                    }
                logger.info("Loaded %d date-slots from sheet '%s'", len(merged), sheet_name)
            except Exception as exc:
                logger.debug("Sheet '%s' not found or empty: %s", sheet_name, exc)

    except Exception as exc:
        logger.error("Failed to load attendance from Google Sheets: %s", exc)

    return merged


# ── Async wrappers (guarded by Lock, offload to thread) ─────────────────────

async def save_to_google_sheets(
    emp_id: str, name: str, dept: str, date_str: str,
    day_name: str, submit_time: str,
    work: str, source: str, username: str,
    is_resubmission: bool = False,
) -> None:
    async with _gs_lock:
        await asyncio.to_thread(
            _save_to_google_sheets_sync,
            emp_id, name, dept, date_str, day_name, submit_time,
            work, source, username, is_resubmission,
        )


async def update_entry_in_google_sheets(emp_id: str, date_str: str, new_work: str) -> None:
    async with _gs_lock:
        await asyncio.to_thread(
            _update_entry_in_google_sheets_sync, emp_id, date_str, new_work,
        )


async def save_leave_to_google_sheets(
    emp_id: str, name: str, dept: str,
    leave_date: str, reason: str, approved_by: str,
    leave_count: int,
) -> None:
    async with _gs_lock:
        await asyncio.to_thread(
            _save_leave_to_google_sheets_sync,
            emp_id, name, dept, leave_date, reason, approved_by, leave_count,
        )


async def load_attendance_from_google_sheets() -> dict:
    """Load history from Google Sheets (async)."""
    return await asyncio.to_thread(_load_attendance_from_google_sheets_sync)


# ── Utility ──────────────────────────────────────────────────────────────────

def get_google_sheet_url() -> str:
    if GOOGLE_SHEET_ID:
        return f"https://docs.google.com/spreadsheets/d/{GOOGLE_SHEET_ID}"
    return ""
